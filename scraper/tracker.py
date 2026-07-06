"""Read/write the flat-hunt tracker (``.xlsx``) with URL-based dedup.

Single ``Flats`` sheet with flat-specific columns (furnishing, balcony/terrace,
size). Coloured priority rows, frozen header, auto-filter, clickable URLs.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .features import OUTDOOR_LABELS
from .models import Listing

FLATS_SHEET = "Flats"
COLS = [
    "Title", "Platform", "URL", "Area", "Postcode", "Price (pcm)", "Bedrooms",
    "Furnishing", "Balcony/Terrace", "Size (sqft)", "Available From",
    "Notes", "Status", "Priority", "Found On",
]
_COL_WIDTHS = {
    "Title": 42, "Platform": 13, "URL": 48, "Area": 24, "Postcode": 10,
    "Price (pcm)": 12, "Bedrooms": 10, "Furnishing": 15, "Balcony/Terrace": 22,
    "Size (sqft)": 11, "Available From": 14, "Notes": 44, "Status": 12,
    "Priority": 10, "Found On": 12,
}

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=11)
_PRIORITY_FILL = {
    "High": PatternFill("solid", fgColor="E2EFDA"),
    "Medium": PatternFill("solid", fgColor="FFFFC7"),
    "Low": PatternFill("solid", fgColor="FCE4D6"),
}
_NEW_STATUS = "NEW 🔴"


def _init_sheet(ws) -> None:
    for i, col in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[cell.column_letter].width = _COL_WIDTHS.get(col, 16)
    ws.freeze_panes = "A2"


def load_or_create(path: str | Path) -> openpyxl.Workbook:
    p = Path(path)
    if p.exists():
        return openpyxl.load_workbook(p)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _init_sheet(wb.create_sheet(FLATS_SHEET))
    return wb


def _values(listing: Listing, today: str) -> dict:
    return {
        "Title": listing.title,
        "Platform": listing.platform,
        "URL": listing.url,
        "Area": listing.area,
        "Postcode": listing.postcode,
        "Price (pcm)": listing.price_pcm,
        "Bedrooms": listing.bed_label or (listing.bed_count if listing.bed_count is not None else ""),
        "Furnishing": listing.furnishing.replace("-", " ").title() if listing.furnishing not in ("", "unknown") else "Unknown",
        "Balcony/Terrace": OUTDOOR_LABELS.get(listing.outdoor, "Not stated"),
        "Size (sqft)": listing.size_sqft,
        "Available From": listing.available_from,
        "Notes": listing.notes,
        "Status": _NEW_STATUS,
        "Priority": listing.priority,
        "Found On": today,
    }


_PRANK = {"High": 0, "Medium": 1, "Low": 2}


def _dupe_key(d: dict):
    """Cross-listing identity: price + bedrooms + postcode. None when postcode is
    missing (too weak to dedup on), so those fall back to URL-only dedup."""
    pc = str(d.get("Postcode") or "").strip().lower()
    if not pc:
        return None
    return (d.get("Price (pcm)"), str(d.get("Bedrooms") or "").strip().lower(), pc)


def _read_rows(ws) -> list[dict]:
    """Read existing data rows as dicts, preserving any manual edits (Status, etc.)."""
    rows: list[dict] = []
    for r in ws.iter_rows(min_row=2):
        if all(c.value in (None, "") for c in r):
            continue
        d = {COLS[i]: (r[i].value if i < len(r) else None) for i in range(len(COLS))}
        url_cell = r[COLS.index("URL")]
        if url_cell.hyperlink:
            d["URL"] = url_cell.hyperlink.target
        rows.append(d)
    return rows


def _write_rows(ws, rows: list[dict]) -> None:
    """Clear the data area and rewrite the given rows with formatting."""
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for d in rows:
        row_idx = ws.max_row + 1
        fill = _PRIORITY_FILL.get(d.get("Priority"))
        for i, col in enumerate(COLS, 1):
            cell = ws.cell(row=row_idx, column=i, value=d.get(col))
            if fill:
                cell.fill = fill
            if col == "URL" and d.get("URL"):
                cell.hyperlink = d["URL"]
                cell.font = Font(color="0563C1", underline="single")
    ws.auto_filter.ref = ws.dimensions


def update_tracker(path: str | Path, listings: list[Listing]) -> dict:
    """Add genuinely-new listings, then sort the sheet by Found On (newest first)
    and priority (High→Low). Dedup is by URL and by (price, beds, postcode) so the
    same flat re-listed or cross-posted to another portal isn't counted as new."""
    wb = load_or_create(path)
    ws = wb[FLATS_SHEET]
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    existing = _read_rows(ws)
    seen_urls = {str(d.get("URL", "")).strip() for d in existing}
    seen_keys = {k for k in (_dupe_key(d) for d in existing) if k}
    new_urls: set[str] = set()
    dupes = 0
    new_rows: list[dict] = []
    for listing in listings:
        row = _values(listing, stamp)
        url = str(row["URL"]).strip()
        key = _dupe_key(row)
        if not url or url in seen_urls or (key and key in seen_keys):
            dupes += 1
            continue
        new_rows.append(row)
        seen_urls.add(url)
        if key:
            seen_keys.add(key)
        new_urls.add(url)

    all_rows = existing + new_rows
    # Found On descending, then priority (High first). reverse=True on the tuple
    # gives newest first and, within a timestamp, High (−0) before Low (−2).
    all_rows.sort(
        key=lambda d: (str(d.get("Found On") or ""), -_PRANK.get(d.get("Priority"), 3)),
        reverse=True,
    )
    _write_rows(ws, all_rows)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return {
        "added": len(new_urls),
        "duplicates": dupes,
        "new_urls": new_urls,
        "rows": all_rows,
    }
